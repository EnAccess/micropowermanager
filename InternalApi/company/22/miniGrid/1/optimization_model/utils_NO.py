from pyomo.environ import DataPortal, Var, value, Expression
import pandas as pd
import numpy as np
import csv
# import xlrd
from openpyxl import load_workbook
import sys
import os.path
from collections import defaultdict

def excel_to_csv(excel_file, csv_file_base_path):
    """Converts excel input data to csv and also return a dictionary of each excel sheet with the data loaded.
    To test out, run utils.py
    """
    df_dict = {}
    workbook = load_workbook(excel_file, data_only=True)
    for sheet_name in workbook.sheetnames:
        # print('processing - ' + sheet_name) # -> N (19.08): To reduce output. I don't think we need this.
        worksheet = workbook.get_sheet_by_name(sheet_name)
        csv_file_full_path = os.path.join(csv_file_base_path, sheet_name.lower().replace(" - ", "_").replace(" ","_") + '.csv')
        with open(csv_file_full_path, 'w') as csvfile:
            writetocsv = csv.writer(csvfile, quoting = csv.QUOTE_ALL)

            for row in worksheet.iter_rows():
                writetocsv.writerow([x.value for x in row]) # list(x.value.encode('utf-8') if type(x.value) == type(b'') else

            # writetocsv.writerow(list(x.value.encode('utf-8') if type(x.value) == type(b'') else x.value for x in worksheet.iter_rows()))
        df_dict[sheet_name.lower().replace(" - ", "_").replace(" ","_")] = pd.read_csv(csv_file_full_path, encoding='utf-8')
        # print(sheet_name + ' has been saved at - ' + csv_file_full_path) # -> N (19.08): To reduce output. I don't think we need this.

    return df_dict

def load_minigrid_data(filename, steps):
    """Loading data from given filename"""

    # Get data file (which is by default stored in the folder "Data")
    filename = os.path.realpath(filename)
    cache_path = os.path.dirname(filename)
    # convert each sheet into a seperate .csv file and load data via pandas
    df = excel_to_csv(filename, cache_path)

    timesteps = int(steps)

    data = {None: {
        'T': {None: [i for i in range(1,timesteps+1)]},
        'tau': {None: float(df['parameters'].iloc[86,3])},
        'D': {i+1: float(df['loada'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)},
        'pv_max': {i+1: float(df['epv'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)},
        'p_load': {i+1: float(df['price_electricity_optional'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)},
        'c_pv': {None: float(df['parameters'].iloc[5,3])},
        'c_bess': {None: float(df['parameters'].iloc[8,3])},
        'E_diesel': {None: float(df['parameters'].iloc[11,3])},
        'c_f':{None: float(df['parameters'].iloc[14,3])},
        'STC_dg': {None: float(df['parameters'].iloc[17,3])},
        'mrt': {None: float(df['parameters'].iloc[20,3])},
        'rt': {None: float(df['parameters'].iloc[23,3])},
        'b_dg': {None: float(df['parameters'].iloc[26,3])},
        'm_dg': {None: float(df['parameters'].iloc[29,3])},
        'q_a': {None: float(df['parameters'].iloc[35,3])},
        # 'q_b': {None: float(df['parameters'].iloc[39,3])},
        'eta_dg': {None: float(df['parameters'].iloc[47,3])},
        'SOC_max': {None: float(df['parameters'].iloc[50,3])},
        'SOC_min': {None: float(df['parameters'].iloc[53,3])},
        'sigma_bess': {None: float(df['parameters'].iloc[56,3])},
        'eta_char': {None: float(df['parameters'].iloc[59,3])},
        'eta_disc': {None: float(df['parameters'].iloc[62,3])},
        'SOC_ini': {None: float(df['parameters'].iloc[65,3])},
        'E_pv_inv': {None: float(df['parameters'].iloc[68,3])},
        'eta_pv_inv': {None: float(df['parameters'].iloc[71,3])},
        'E_ir': {None: float(df['parameters'].iloc[74,3])},
        'eta_rect': {None: float(df['parameters'].iloc[77,3])},
        'eta_bess_inv': {None: float(df['parameters'].iloc[80,3])},
        'lower_limit_diesel': {None: float(df['parameters'].iloc[83,3])},
        'L_char': {None: float(df['parameters'].iloc[89,3])},
        'L_disch': {None: float(df['parameters'].iloc[92,3])},
        'p_mi': {None: float(df['parameters'].iloc[95,3])},
        'crypto_max': {None: float(df['parameters'].iloc[98,3])},
        'm_mi': {None: float(df['parameters'].iloc[101,3])},
        'b_mi': {None: float(df['parameters'].iloc[104,3])},
        'K_mi': {None: float(df['parameters'].iloc[107,3])},
        'crypto_min': {None: float(df['parameters'].iloc[110,3])}
    }}
    print(data)
    return data


def build_dict_list(indexes, variables):
    """Build a dictionary with unique indexes and a list of variables with the same key"""
    relations = defaultdict(list)
    for idx, variable in zip(indexes, variables):
        relations[idx].append(variable)
    return relations

def build_dataframes_from_pages(pages):
    """Each key of the page is a dataframe. Every variable is a column and the indexes the row index"""
    df_dict = dict()
    for set_name, variable_list in pages.items():
        index = pd.Index(variable_list[0].index_set(), name=set_name)
        columns = [v.name for v in variable_list]
        data = np.array([[x() for x in v.values()] for v in variable_list]).T
        df = pd.DataFrame(data=data, columns=columns, index=index)
        df_dict[set_name] = df
    return df_dict

def df_dict_to_excel(filename, df_dict_var, df_dict_expr, df_dict_obj, data):
    # Load workbook to be written into:
    book = load_workbook(filename)
    # Define ExcelWriter:
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    # Make sure writer doesn't erase existing sheets, but only modifies them:
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    """
    N (19.08): For the writing of variables and expressions, we are relying on the fact that the dataframe that was built keeps the order of the variables and expressions.
    We should add some asserts to make sure they're written to the excel in the right order.
    """
    # Write variables dataframe into specified sheet:
    for index, df in df_dict_var.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Mini-grid results', startrow=1, startcol=1)

    # Write expressions dataframe into specified sheet:
    for index, df in df_dict_expr.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Objective Cost', startrow=1, startcol=2)

    # Write objective into specified sheet:
    for index, df in df_dict_obj.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Objective Cost', startrow=1, startcol=0)

    # Write parameters into specified sheet:
    param_sheet = book['Inputs-Parameters']
    param_names = ['SOC_max', 'eta_pv_inv',	'eta_disc', 'eta_bess_inv', 'eta_rect', 'eta_char', 'D', 'p_load', 'pv_max']
    for i in range(6):
        param_sheet.cell(column=i+1,row=2, value=data[None][param_names[i]][None])
    for i in range(6,9):
        for j in range(1,len(data[None][param_names[i]])+1):
            param_sheet.cell(column=i+2, row=j+1, value=data[None][param_names[i]][j])

    # Save the file:
    writer.save()

def dump_minigrid_data(filename, instance, data):
    """Load data from the instance and store it in an excel file"""
    filename = os.path.realpath(filename)

    # Build DataFrame of all variables:
    variables = list(instance.component_objects(ctype=Var))
    names = [v.getname() for v in variables]
    indexes= [v.index_set().name for v in variables]
    pages = build_dict_list(indexes, variables)
    df_dict_vars = build_dataframes_from_pages(pages)

    # Build DataFrame of all expressions:
    expressions = list(instance.component_objects(ctype=Expression))
    names = [expr.getname() for expr in expressions]
    indexes= [expr.index_set().name for expr in expressions]
    pages = build_dict_list(indexes, expressions)
    df_dict_expr = build_dataframes_from_pages(pages)

    # Build dataframe of objective value:
    df_dict_obj = pd.DataFrame({None:{'Objective': value(instance.Obj)}})

    # Writing all dataframes into excel file
    df_dict_to_excel(filename, df_dict_vars, df_dict_expr, df_dict_obj, data)

    print("")
    print('### The results were successfully written into ', filename)

from pprint import pprint

# For testing:
if __name__ == '__main__':
    # Test excel_to_csv
    cwd = os.path.dirname(os.path.realpath(__file__))
    path = str(os.path.join(cwd, "Data", 'Data_input_Inensus.xlsm'))
    # data = excel_to_csv(path, os.path.join(cwd, "Data"))
    dirname = os.path.dirname(__file__)
    filename_input = os.path.join(dirname, 'Data', 'Data_input_Inensus.xlsm')
    data = load_minigrid_data(filename_input, 96)
    pprint(data)
