from pyomo.environ import DataPortal, Var, value, Expression
import pandas as pd
import numpy as np
import csv
import xlrd
from openpyxl import load_workbook
import sys
import os.path
from collections import defaultdict

def excel_to_csv(excel_file, csv_file_base_path):
    """Converts excel input data to csv and also return a dictionary of each excel sheet with the data loaded.

    To test out, run utils.py


    """
    df_dict = {}
    workbook = load_workbook(excel_file, data_only=True)
    for sheet_name in workbook.sheetnames:
        # print('processing - ' + sheet_name) # -> N (19.08): To reduce output. I don't think we need this.
        worksheet = workbook[sheet_name]
        csv_file_full_path = os.path.join(csv_file_base_path, sheet_name.lower().replace(" - ", "_").replace(" ","_") + '.csv')
        with open(csv_file_full_path, 'w') as csvfile:
            writetocsv = csv.writer(csvfile, quoting = csv.QUOTE_ALL)
            for row in worksheet.iter_rows():
                print(row)
                a = list(x.encode('utf-8') if type(x) == type(b'') else x.internal_value for x in row
                    )
                writetocsv.writerow(a

                )
        df_dict[sheet_name.lower().replace(" - ", "_").replace(" ","_")] = pd.read_csv(csv_file_full_path, encoding='utf-8')
        # print(sheet_name + ' has been saved at - ' + csv_file_full_path) # -> N (19.08): To reduce output. I don't think we need this.
    return df_dict

def meta_to_dict(data):
    meta_data = data['meta_data']
    return {'Creation Date': meta_data.columns[2], 'Timezone': meta_data.iloc[0, 2], 'Timestamp': meta_data.iloc[1,2]}



def load_minigrid_data(filename, num_days):
    """Loading data from given filename"""

    # Get data file (which is by default stored in the folder "Data")
    filename = os.path.realpath(filename)
    cache_path = os.path.dirname(filename)
    # convert each sheet into a seperate .csv file and load data via pandas
    df = excel_to_csv(filename, cache_path)

    # Calculate time steps out of the given number of days:
    timesteps = int(num_days) * 96

    data = {None: {
        'T': {None: [i for i in range(1,timesteps+1)]},
        'D': {i+1: float(df['loada'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)}, # import and convert to kW
        'pv_max': {i+1: float(df['epv'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)},
        'p_load': {i+1: float(df['price_electricity_optional'].iloc[1:timesteps+1,1].to_numpy()[i]) for i in range(timesteps)}, # TODO confirm this
        'c_pv': {None: df['parameters'].iloc[6,3]},
        'c_bess': {None: df['parameters'].iloc[9,3]},
        'E_diesel': {None: df['parameters'].iloc[12,3]},
        'c_f':{None: df['parameters'].iloc[15,3]},
        'STC_dg': {None: df['parameters'].iloc[18,3]},
        'mrt': {None: df['parameters'].iloc[21,3]},
        'rt': {None: df['parameters'].iloc[24,3]},
        'b_dg': {None: df['parameters'].iloc[27,3]},
        'm_dg': {None: df['parameters'].iloc[30,3]},
        'c_nse': {None: df['parameters'].iloc[33,3]},
        # 'p_load': {None: df['parameters'].iloc[42,3]},
        'eta_dg': {None: df['parameters'].iloc[52,3]},
        'SOC_max': {None: df['parameters'].iloc[55,3]},
        'SOC_min': {None: df['parameters'].iloc[58,3]},
        'sigma_bess': {None: df['parameters'].iloc[61,3]},
        'L_disc': {None: df['parameters'].iloc[67,3]},
        'L_char': {None: df['parameters'].iloc[64,3]},
        'eta_char': {None: df['parameters'].iloc[70,3]},
        'eta_disc': {None: df['parameters'].iloc[73,3]},
        'SOC_ini': {None: df['parameters'].iloc[76,3]},
        'E_pv_inv': {None: df['parameters'].iloc[79,3]},
        'eta_pv_inv': {None: df['parameters'].iloc[82,3]},
        'E_ir': {None: df['parameters'].iloc[85,3]},
        'eta_bess_inv': {None: df['parameters'].iloc[91,3]},
        'eta_rect': {None: df['parameters'].iloc[88,3]}

    },
    'Meta_Data': meta_to_dict(df)}
    return data


def get_dataportal(model):
    """
    Incomplete yet
    -> N (19.08): What is this function for? Do you want to use DataPortal? As said before, we don't need this.
                  DataPortal() is no longer under development. Using pandas and Dataframes works perfectly.
    --> T: Ok let's remove it
    """
    data = DataPortal()
    data.load(filename='data/sets.yaml')
    return data

def build_dict_list(indexes, variables):
    """Build a dictionary with unique indexes and a list of variables with the same key"""
    relations = defaultdict(list)
    for idx, variable in zip(indexes, variables):
        relations[idx].append(variable)
    return relations

def build_dataframes_from_pages(pages):
    """Each key of the page is a dataframe. Every variable is a column and the indexes the row index"""
    df_dict = dict()
    for set_name, variable_list in pages.items():
        index = pd.Index(variable_list[0].index_set(), name=set_name)
        columns = [v.name for v in variable_list]
        data = np.array([[x() for x in v.values()] for v in variable_list]).T
        df = pd.DataFrame(data=data, columns=columns, index=index)
        df_dict[set_name] = df
    return df_dict

def df_dict_to_excel(filename, df_dict_var, df_dict_expr, df_dict_obj, data):
    # Load workbook to be written into:
    book = load_workbook(filename)
    # Define ExcelWriter:
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    # Make sure writer doesn't erase existing sheets, but only modifies them:
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    """
    For the writing of variables and expressions, we are relying on the fact that the dataframe that was built keeps the order of the variables and expressions.
    We should add some asserts to make sure they're written to the excel in the right order.
    """
    # Write variables dataframe into specified sheet:
    for index, df in df_dict_var.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Mini-grid results', startrow=1, startcol=1)

    # Write expressions dataframe into specified sheet:
    for index, df in df_dict_expr.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Objective Cost', startrow=1, startcol=2)

    # Write variables dataframe into specified sheet:
    for index, df in df_dict_obj.items():
        df.to_excel(writer, header=False, index=False, sheet_name='Objective Cost', startrow=1, startcol=0)

    # Write parameters into specified sheet:
    param_sheet = book['Inputs-Parameters']
    param_names = ['SOC_max', 'eta_pv_inv',	'eta_disc', 'eta_bess_inv', 'eta_rect', 'eta_char', 'D', 'p_load', 'pv_max']
    for i in range(6):
        param_sheet.cell(column=i+1,row=2, value=data[None][param_names[i]][None])
    for i in range(6,9):
        for j in range(1,len(data[None][param_names[i]])+1):
            param_sheet.cell(column=i+2, row=j+1, value=data[None][param_names[i]][j])

    if 'Meta Data' not in book.sheetnames:
        book.create_sheet('Meta Data')
    sheet = book['Meta Data']
    row = 1
    for key in data['Meta_Data'].keys():
        sheet.cell(row, 1, value = key)
        sheet.cell(row, 3, value=data['Meta_Data'].get(key))
        row +=1
    # Save the file:
    writer.save()

def dump_minigrid_data(filename, instance, data): # TODO: Test with different indexes -> N (19.08) : What do you mean ?
    """Load data from the instance and store it in an excel file"""
    filename = os.path.realpath(filename)

    # Build DataFrame of all variables:
    variables = list(instance.component_objects(ctype=Var))
    names = [v.getname() for v in variables]
    indexes= [v.index_set().name for v in variables]
    pages = build_dict_list(indexes, variables)
    df_dict_vars = build_dataframes_from_pages(pages)

    # Build DataFrame of all expressions:
    expressions = list(instance.component_objects(ctype=Expression))
    names = [expr.getname() for expr in expressions]
    indexes= [expr.index_set().name for expr in expressions]
    pages = build_dict_list(indexes, expressions)
    df_dict_expr = build_dataframes_from_pages(pages)

    # Build dataframe of objective value:
    df_dict_obj = pd.DataFrame({None:{'Objective': value(instance.Obj)}})

    # Writing all dataframes into excel file
    df_dict_to_excel(filename, df_dict_vars, df_dict_expr, df_dict_obj, data)

    print("")
    print('### The results were successfully written into ', filename)

# N (19.08): we can delete this:
if __name__ == '__main__':
    # Test excel_to_csv
    cwd = os.path.dirname(os.path.realpath(__file__))
    path = str(os.path.join(cwd, "Data", 'Data_input_Inensus.xlsm'))
    excel_to_csv(path, os.path.join(cwd, "Data"))
