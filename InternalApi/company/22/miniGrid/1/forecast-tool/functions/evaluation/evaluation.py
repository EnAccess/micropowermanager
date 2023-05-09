import os

import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_absolute_percentage_error

from functions.utils.path_utils import check_path_exist
from communication.socket_client import SocketClient

import logging

module_logger = logging.getLogger('Forecasting-Tool.forecasting_module')


def calculate_metrics(eval_data, measurements, date_time):
    """
    calculate evaluation metrics

    Parameter
    ---------
    eval_data : pd.DataFrame
        data to calculate metrics for
    measurements : pd.DataFrame
        recent measurements
    date_time : str
        date_time string when prediction was made

    Return
    ------
    results : pd.DataFrame
        evaluation results

    """
    results = pd.DataFrame()
    data = eval_data.copy(deep=True)
    data['measurements'] = measurements
    data = data[(data['measurements'] == data['measurements'])
                & (data['measurements'] > 0)]
    measured_energy = data['measurements'].sum()
    data = data.fillna(0)
    if not data.empty:
        for col in eval_data.columns:
            results.loc[date_time, f'mae_{col}'] = mean_absolute_error(y_true=data['measurements'].values,
                                                                       y_pred=data[col].values)
            results.loc[date_time, f'mse_{col}'] = mean_squared_error(y_true=data['measurements'].values,
                                                                      y_pred=data[col].values)
            try:
                results.loc[date_time, f'mape_{col}'] = mean_absolute_percentage_error(
                    y_true=data['measurements'].values, y_pred=data[col].values)
            except:
                pass
            results.loc[date_time, f'energy_diff_{col}'] = abs(
                eval_data[col].sum() - measured_energy.sum())
    else:
        pass
    return results


class ForecastEvaluation:
    """
    Class to evaluate forecasts and gives recommendations to train

    Methods
    -------
    __read_recent_forecasts()
        reading in all forecasts which were made
    evaluate_recent_forecasts
        evaluates forecasts if data is available
    __eval_ml_retrain
        checks if ml algorithms should be retrained
    __eval_ml_reset
        checks if ml algorithms need a reset
    __eval_metrics
        calculates metrices
    recommend
        recommends retrain, reset or to do nothing
    write_to_excel
        writes predictions to excel for optimizer
    write_to_csv
        saves prediction as csv

    Attributes
    ----------
    recent_forecasts : dict with pd.DataFrame
        calculated metrics of recent forecasts
    forecast_horizon : dict with strings
        forecast horizons of long and short term forecasts
    frequencies : dict with strings
        data frequencies of long and short term forecasts
    recommendation : dict
        collected recommendations to retrain, reset or nothing
    sheet : dict
        which excel sheet to put data in for optimizer
    excel_path : str
        path to excel file to put predictions in

    """

    def __init__(self, folder):
        """
        Attributes
        ----------
        recent_forecasts : dict with pd.DataFrame
            calculated metrics of recent forecasts
        forecast_horizon : dict with strings
            forecast horizons of long and short term forecasts
        frequencies : dict with strings
            data frequencies of long and short term forecasts
        recommendation : dict
            collected recommendations to retrain, reset or nothing
        sheet : dict
            which excel sheet to put data in for optimizer
        excel_path : str
            path to excel file to put predictions in

        """
        self.folder = folder
        self.recent_forecasts = {'pv_short': pd.DataFrame(), 'pv_long': pd.DataFrame(), 'lf_short': pd.DataFrame(),
                                 'lf_long': pd.DataFrame()}
        self.forecast_horizon = {
            'pv_short': '2D', 'pv_long': '90D', 'lf_short': '2D', 'lf_long': '90D'}
        self.frequencies = {'pv_short': '15Min',
                            'pv_long': '1H', 'lf_short': '15Min', 'lf_long': '1H'}
        self.recommendation = {'pv_short': None,
                               'pv_long': None, 'lf_short': None, 'lf_long': None}
        self.sheet = {'pv_short': 'EPV', 'pv_long': None,
                      'lf_short': 'loadA', 'lf_long': None}
        self.excel_path = folder + '/resources/05_output/raw/excel_sheet.xlsx'
        os.makedirs(
            folder + '/resources/05_output/saved_predictions/', exist_ok=True)

    def __read_recent_forecasts(self, path, time_now, forecast_horizon, current_data, freq, recent_forecast_metrics):
        """
        reads in forecasts done

        Parameter
        ---------
        path : str
            path to forecasts
        time_now : datetime
            current timestamp
        forecast_horizon : str
            forecast horizon for part
        current_data : pd.DataFrame
            all measurements up to timestamp
        freq : str
            frequency of data
        recent_forecast_metrics : pd.DataFrame
            all evaluated forecasts

        Return
        ------
        recent_forecast_metrics : pd.DataFrame
            evaluated forecasts with new added metrics

        """
        current_data = current_data.copy(deep=True).resample(freq).mean()
        for subdir, dirs, files in os.walk(path):
            for file in files:
                date_time_split = file.split(' ')
                date, time = date_time_split[0], date_time_split[1].split(
                    '+')[0].replace('-', ':')
                date_time = pd.to_datetime(
                    date + ' ' + time).tz_localize('Africa/Dar_es_Salaam')
                if date_time >= time_now - pd.to_timedelta(
                        forecast_horizon) or date_time > current_data.last_valid_index():
                    pass
                else:
                    if not date_time in recent_forecast_metrics.index:
                        walk_path = os.path.join(subdir, file)
                        data = pd.read_csv(walk_path)
                        data.index = pd.to_datetime(data['Unnamed: 0'])
                        data = data.drop(columns=['Unnamed: 0'])
                        results = calculate_metrics(
                            data, current_data, date_time)
                        if recent_forecast_metrics is None:
                            recent_forecast_metrics = results
                        else:
                            recent_forecast_metrics = pd.concat(
                                [recent_forecast_metrics, results])
                    else:
                        pass
        return recent_forecast_metrics

    def evaluate_recent_forecasts(self, time_now, part, current_data):
        """
        Function

            evaluate forecasts for which data is available

        Parameter

            time_now : datetime
                current timestamps

            part : str
                part to evaluate

            current_data : pd.DataFrame
                all measurements available

        Return

            None

        """
        module_logger.info(f'Evaluating recent forecasts for: {part}')
        path = self.folder + f'/resources/03_predictions/{part}/'
        freq = self.frequencies.get(part)
        forecast_horizon = self.forecast_horizon.get(part)
        path_exists = check_path_exist(path)
        recent_forecast_metrics = self.recent_forecasts.get(part)
        if path_exists is True:
            recent_forecast_metrics = self.__read_recent_forecasts(path, time_now, forecast_horizon, current_data, freq,
                                                                   recent_forecast_metrics)
            self.recent_forecasts[part] = recent_forecast_metrics
            if not recent_forecast_metrics.empty:
                metric_path = self.folder + f'/resources/04_evaluation/{part}/'
                check_path_exist(metric_path)
                recent_forecast_metrics = recent_forecast_metrics[
                    recent_forecast_metrics.index >= recent_forecast_metrics.last_valid_index() - pd.to_timedelta(
                        '40D')]
                recent_forecast_metrics.to_csv(metric_path + 'metrics.csv')
        else:
            pass
        module_logger.info(
            f'Evaluation of recent forecasts for: {part} completed')

    def __eval_ml_retrain(self, metric_frame, eval_time='7D'):
        """
        Evaluates if ML-Algorithm needs Retraining

        Parameter
        ---------
        metric_frame : pd.DataFrame
            all calculated metrics up to now
        eval_time : str
            look back of evaluation

        Result
        -------
        return is bool:
            True = Retrain
            False = No Training
        """
        mase = metric_frame['mae_power'] / metric_frame['mae_persistency']
        recent_mase = mase[mase.index >= metric_frame.last_valid_index(
        ) - pd.to_timedelta(eval_time)]
        if len(recent_mase[recent_mase >= 1]) / len(
            recent_mase) >= 0.5 and metric_frame.first_valid_index() <= metric_frame.last_valid_index() - pd.to_timedelta(
                '4D'):
            return True
        else:
            return False

    def __eval_ml_reset(self, metric_frame, eval_time='30D'):
        """
        Evaluates if ML-Algorithm needs Reset

        Parameter
        ---------
        metric_frame : pd.DataFrame
            all calculated metrics up to now
        eval_time : str
            look back of evaluation

        Result
        -------
        return is bool:
            True = Reset
            False = No Reset
        """
        mase = metric_frame['mae_power'] / metric_frame['mae_persistency']
        recent_mase = mase[mase.index >= metric_frame.last_valid_index(
        ) - pd.to_timedelta(eval_time)]
        if len(recent_mase[recent_mase >= 1]) / len(
            recent_mase) >= 0.5 and metric_frame.first_valid_index() <= metric_frame.last_valid_index() - pd.to_timedelta(
                '14D'):
            return True
        else:
            return False

    def __eval_metrics(self, metric_frame):
        """
        get recommendation which algorithm to use

        Parameter
        ---------
        metric_frame : pd.DataFrame
            all evaluated metrics

        Return
        ------
        recommendation based on metric
        """
        energy_col = [
            col for col in metric_frame.columns if 'energy_diff' in col]
        energy_metrics = metric_frame[energy_col].mean(axis=0)
        recent_minimum = energy_metrics.min()
        recommendation = energy_metrics[energy_metrics ==
                                        recent_minimum].index.format()
        return recommendation[0].replace('energy_diff_', '')

    def recommend(self):
        """
        Recommends to retrain, resets or nothing

        Parameter
        ---------
        None

        Return
        ------
        Recommendation

        """
        for key in self.recent_forecasts.keys():
            data = self.recent_forecasts.get(key)
            recommendation_dict = {}
            if data is not None and not data.empty:
                recommendation_dict['prediction'] = self.__eval_metrics(data)
                recommendation_dict['retrain'] = self.__eval_ml_retrain(data)
                recommendation_dict['reset'] = self.__eval_ml_reset(data)
            else:
                recommendation_dict['prediction'] = 'persistency'
                recommendation_dict['retrain'] = False
                recommendation_dict['reset'] = False
            self.recommendation[key] = recommendation_dict
            module_logger.info(
                f'Recommendation for {key}: Training: {recommendation_dict.get("retrain")}, Reset: {recommendation_dict.get("reset")}')

    def write_to_excel(self, forecasts, time_now):
        """
        Writes predictions to excel for optimization

        Parameter
        ---------
        forecast : pd.DataFrame
            all forecasts for part

        Return
        ------
        None
        """
        module_logger.info('Writing forecasts to EXCEL-File')
        book = load_workbook(self.excel_path, read_only=False,
                             keep_vba=True, keep_links=True, data_only=True)
        for key in self.recommendation.keys():
            if self.sheet.get(key) is not None:
                sheet = book[self.sheet.get(key)]
                current_recommendation = self.recommendation.get(key)
                forecast = forecasts.get(key)
                data_of_the_column = forecast[current_recommendation.get(
                    'prediction')]
                data_of_one_day = data_of_the_column.reset_index()
                data_of_one_day.index = data_of_one_day.index + 1
                for row in range(0, len(data_of_one_day)):
                    value = data_of_one_day.iloc[row, 1]
                    sheet.cell(row=row + 3, column=2, value=value)
        if 'Meta Data' not in book.sheetnames:
            book.create_sheet('Meta Data')
        sheet = book['Meta Data']
        sheet.cell(row=1, column=1, value='Creation Date')
        sheet.cell(row=3, column=1, value='Timestamp')
        sheet.cell(row=2, column=1, value='Timezone')
        sheet.cell(row=2, column=3, value='UTC')
        sheet.cell(row=1, column=3, value=time_now)
        sheet.cell(row=3, column=3, value=int(time_now.timestamp()))
        output_path = self.folder + '/resources/05_output/predictions.xlsm'
        book.save(output_path)
        module_logger.info('Forecast written to EXCEL-File')

    def write_to_csv(self, forecasts, time_now):
        """
        Writes predictions to csv to store for further evaluation

        Parameter
        ---------
        forecasts : pd.DataFrame
            all forecasts for part
        time_now : datetime
            current timestamp

        Return
        ------
        None
        """
        module_logger.info('Writing Forecasts to CSV')
        recommended_data = pd.DataFrame()
        for key in self.recommendation.keys():
            if self.sheet.get(key) is not None:
                current_recommendation = self.recommendation.get(key)
                forecast = forecasts.get(key)
                data_of_the_column = forecast[current_recommendation.get(
                    'prediction')]
                recommended_data[key] = data_of_the_column
        time_now_cleared = str(time_now).replace(':', '-')

        recommended_data.to_csv(
            self.folder + f'/resources/05_output/saved_predictions/{time_now_cleared}.csv')
        module_logger.info('Forecast written to CSV-Files')
